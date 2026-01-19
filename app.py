"""
🚀 OTOMASI SUMMARY EVENT PROMO v5.1
Streamlit Web Application
Revisi: Pembersihan Tanggal Agresif, Fix Sheet Skipping, & Periode Extraction
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ============================================================
# KONFIGURASI HALAMAN
# ============================================================
st.set_page_config(
    page_title="Otomasi Summary Promo",
    page_icon="📊",
    layout="wide"
)

# ============================================================
# FUNGSI-FUNGSI UTAMA (REVISED)
# ============================================================

def clean_mekanisme(text):
    """
    Membersihkan angka di awal mekanisme.
    Contoh: '1. BELI 1 KARTON' -> 'BELI 1 KARTON'
    """
    if not isinstance(text, str):
        return text
    # Hapus angka diikuti titik atau kurung di awal string (misal: "1.", "1)", "1 ")
    return re.sub(r'^\s*\d+[\.\)]\s*', '', text).strip()

def clean_promo_name(nama_promo):
    """
    Membersihkan tanggal dari nama promo secara agresif.
    Menangani format: '20-26 AUG 25', '1-31 JANUARI 2026', dsb.
    """
    if not nama_promo:
        return nama_promo
    
    # Regex pattern untuk berbagai format tanggal
    bulan_ind = r'(?:JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER|JAN|FEB|MAR|APR|MEI|JUN|JUL|AGU|SEP|OKT|NOV|DES)'
    bulan_eng = r'(?:JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER|JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)'
    bulan = f'(?:{bulan_ind}|{bulan_eng})'
    
    date_patterns = [
        # Format: 20-26 AUG 25 (Tanggal pendek, tahun 2 digit)
        rf'\d{{1,2}}\s*[-–]\s*\d{{1,2}}\s+{bulan}\s*\d{{2,4}}',
        # Format: 1-31 JANUARI 2026
        rf'\d{{1,2}}\s*[-–]\s*\d{{1,2}}\s+{bulan}\s*\d{{4}}',
        # Format: 1 JAN - 1 FEB 2026
        rf'\d{{1,2}}\s+{bulan}\s*[-–]\s*\d{{1,2}}\s+{bulan}\s*\d{{2,4}}',
        # Format: PERIODE 1-31 JAN
        rf'PERIODE\s*:?\s*\d{{1,2}}.*?\d{{4}}',
        # Tanggal di akhir string: ... AUG 25
        rf'\s+\d{{1,2}}\s+{bulan}\s*\d{{2,4}}$',
        # Format slash: 20/08/25
        r'\d{{1,2}}/\d{{1,2}}/\d{{2,4}}'
    ]
    
    result = nama_promo
    for pattern in date_patterns:
        result = re.sub(pattern, '', result, flags=re.IGNORECASE)
    
    # Bersihkan sisa-sisa karakter aneh di ujung
    result = re.sub(r'[\(\)\.,\-\s]+$', '', result).strip()
    return result

def extract_nama_promo(text):
    if not text or not isinstance(text, str):
        return ""

    s = text.upper().strip()

    # 1. Coba pola "Angka - Nama" (Contoh: "382 - PROMO X")
    m = re.search(r'\b\d+\s*[-–]\s*(.+)', s)
    if m:
        s = m.group(1)
    else:
        # REVISI: Jika tidak ada angka di depan, tapi mengandung kata kunci, ambil seluruhnya
        # Ini mengatasi masalah sheet yang terlewat karena tidak punya nomor urut
        if not any(x in s for x in ["PROMO", "DISKON", "POTONGAN", "PROGRAM"]):
            # Jika tidak mirip nama promo, kembalikan kosong agar diproses fallback lain
            # Tapi jika string cukup panjang dan ada di header, mungkin itu nama promonya
            if len(s) < 5: 
                return ""

    # Bersihkan tanggal
    s = clean_promo_name(s)
    
    # Hapus metadata dalam kurung (misal: "(KVI)") jika diinginkan, 
    # tapi kadang ini info penting. Jika ingin dihapus, uncomment baris bawah:
    # s = re.sub(r'\(.*?\)', '', s)

    return s.strip()

def extract_periode(text):
    """
    Ekstraksi periode dengan regex yang lebih fleksibel
    """
    if not isinstance(text, str): return ""
    
    bulan = r'(?:JAN|FEB|MAR|APR|MEI|JUN|JUL|AGU|SEP|OKT|NOV|DES|AUG|OCT|DEC|[A-Z]{3,})'
    
    # Pola: 20-26 AUG 25 atau 01 JAN - 31 JAN 2026
    pattern = rf'(\d{{1,2}}\s*[-–]\s*\d{{1,2}}\s+{bulan}\s*\d{{2,4}}|\d{{1,2}}\s+{bulan}\s*[-–]\s*\d{{1,2}}\s+{bulan}\s*\d{{2,4}})'
    
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return ""

def extract_promo_info_flexible(df):
    nama_promo = ""
    periode_text = ""
    mekanisme = ""

    # Loop 6 baris pertama untuk mencari info header
    for row_idx in range(min(8, len(df))):
        for col_idx in range(min(4, len(df.columns))):
            cell_value = df.iloc[row_idx, col_idx]
            if pd.isna(cell_value):
                continue

            cell_str = str(cell_value).strip()
            
            # --- CARI NAMA PROMO ---
            if not nama_promo:
                extracted = extract_nama_promo(cell_str)
                # Kriteria: Mengandung "PROMO" atau panjangnya cukup logis dan bukan tanggal doang
                if extracted and len(extracted) > 5 and not re.match(r'^\d+$', extracted):
                    nama_promo = extracted
                    # Jangan continue, karena di cell yang sama mungkin ada periode (jarang sih)
            
            # --- CARI PERIODE ---
            if not periode_text:
                # Cek apakah cell ini khusus periode?
                if "PERIODE" in cell_str.upper():
                    # Ambil isinya
                    cleaned_periode = cell_str.upper().replace("PERIODE", "").replace(":", "").strip()
                    periode_text = cleaned_periode
                else:
                    # Coba cari pola tanggal di string
                    found_periode = extract_periode(cell_str)
                    if found_periode:
                        periode_text = found_periode

            # --- CARI MEKANISME ---
            # Mekanisme biasanya ada di baris > 0
            if row_idx > 0 and not mekanisme:
                # Kata kunci mekanisme
                keywords = ['beli', 'buy', 'min', 'gratis', 'free', 'disc', 'potongan', 'bonus', 'cashback', 'dapat', 'setiap']
                if any(kw in cell_str.lower() for kw in keywords):
                    mekanisme = clean_mekanisme(cell_str) # REVISI: Bersihkan angka 1.

    # Fallback Mekanisme: Jika regex keyword gagal, ambil baris ke-3 kolom pertama (asumsi posisi standar)
    if not mekanisme and len(df) > 2:
        val = str(df.iloc[2, 0]).strip()
        if len(val) > 5:
            mekanisme = clean_mekanisme(val)

    return nama_promo, periode_text, mekanisme
    
def find_header_row(df):
    """
    Mencari row yang berisi header kolom data
    """
    header_keywords = ['no', 'customer', 'count', 'claim', 'sales', 'amount', 'left', 'bonus', 'qty', 'total', 'sisa']
    
    for row_idx in range(min(12, len(df))): # Scan sampai baris 12
        try:
            row_values = [str(v).lower().strip() for v in df.iloc[row_idx] if pd.notna(v)]
            matches = sum(1 for kw in header_keywords if any(kw in val for val in row_values))
            if matches >= 2: # Cukup 2 match untuk lebih longgar
                return row_idx
        except Exception:
            continue
    
    return None

def find_summary_row(df, header_row):
    """
    Mencari row summary/total
    """
    if header_row is None:
        return 6 if len(df) > 6 else None
    
    # Cari di bawah header
    start_search = header_row + 1
    # Batasi pencarian max 10 baris ke bawah atau sampai habis
    end_search = min(start_search + 15, len(df))
    
    for row_idx in range(start_search, end_search):
        try:
            row_values = df.iloc[row_idx]
            first_val = str(row_values.iloc[0]).lower() if pd.notna(row_values.iloc[0]) else ""
            
            # Ciri-ciri baris summary
            if 'all' in first_val or 'total' in first_val or 'grand' in first_val:
                return row_idx
            
            # Atau baris pertama yang punya banyak angka (jika tidak ada label Total)
            numeric_count = sum(1 for v in row_values if pd.notna(v) and isinstance(v, (int, float)))
            # Jika > 3 kolom berisi angka, kemungkinan ini baris total
            if numeric_count >= 3:
                return row_idx
        except Exception:
            continue
    
    return None # Biarkan None jika tidak ketemu, nanti ditangani process_sheet

def find_column_by_keywords(df, header_row, keywords):
    if header_row is None or header_row >= len(df):
        return None
    
    try:
        # Cek di header_row dan header_row+1 (antisipasi merge cell)
        rows_to_check = [header_row]
        if header_row + 1 < len(df):
            rows_to_check.append(header_row + 1)

        for check_row in rows_to_check:
            for col_idx in range(len(df.columns)):
                cell_value = df.iloc[check_row, col_idx]
                if pd.isna(cell_value): continue
                
                cell_str = str(cell_value).lower().strip()
                if any(kw in cell_str for kw in keywords):
                    return col_idx
    except Exception:
        pass
    
    return None

def safe_convert_number(value):
    if pd.isna(value): return None
    try:
        if isinstance(value, str):
            # Hapus karakter non-numerik kecuali titik dan minus
            value = re.sub(r'[^\d.\-]', '', value.replace(',', ''))
        num = float(value)
        return num if not np.isnan(num) else None
    except (ValueError, TypeError):
        return None

def process_sheet_robust(df, sheet_name):
    result = {
        'Nama Promo': '',
        'Mekanisme Promo': '',
        'Periode Promo': '',
        'All Count': None,
        'All Claim': None,
        'Sales Amount': None,
        'Amount': None,
        'Left': None
    }
    
    try:
        # REVISI: Jangan skip sheet hanya karena baris sedikit, kadang summary ada di sheet kecil
        if len(df) < 2:
            return None, f"⏭️ Sheet '{sheet_name}': Kosong"
        
        # 1. Extract info teks (Nama, Periode, Mekanisme)
        nama_promo, periode_text, mekanisme = extract_promo_info_flexible(df)
        
        if not nama_promo:
            # Last resort: Gunakan nama sheet sebagai nama promo jika di konten tidak ketemu
            # Ini membantu menaikkan jumlah sheet yang terbaca
            nama_promo = sheet_name
        
        result['Nama Promo'] = nama_promo
        result['Mekanisme Promo'] = mekanisme
        result['Periode Promo'] = periode_text
        
        # 2. Cari Header & Data
        header_row = find_header_row(df)
        
        # Jika header tidak ketemu, tapi ada data numerik di baris-baris tertentu, coba tebak
        # Tapi idealnya kita butuh header row.
        
        summary_row_idx = find_summary_row(df, header_row)
        
        if summary_row_idx is not None:
            summary_row = df.iloc[summary_row_idx]
            num_cols = len(df.columns)
            
            # --- MAPPING KOLOM OTOMATIS ---
            count_col = find_column_by_keywords(df, header_row, ['count', 'jumlah customer', 'total customer', 'cust'])
            claim_col = find_column_by_keywords(df, header_row, ['claim', 'klaim', 'redeem'])
            sales_col = find_column_by_keywords(df, header_row, ['sales amount', 'sales amt', 'nilai penjualan', 'sales'])
            amount_col = find_column_by_keywords(df, header_row, ['amount', 'bonus amount', 'nilai bonus', 'bonus', 'amt'])
            left_col = find_column_by_keywords(df, header_row, ['left', 'sisa', 'remaining', 'stock'])
            
            # Fallback posisi kolom (hardcoded index) jika deteksi gagal
            # Disesuaikan dengan asumsi format umum
            if count_col is not None: result['All Count'] = safe_convert_number(summary_row.iloc[count_col])
            elif num_cols > 3: result['All Count'] = safe_convert_number(summary_row.iloc[3])
            
            if claim_col is not None: result['All Claim'] = safe_convert_number(summary_row.iloc[claim_col])
            elif num_cols > 4: result['All Claim'] = safe_convert_number(summary_row.iloc[4])
            
            if sales_col is not None: result['Sales Amount'] = safe_convert_number(summary_row.iloc[sales_col])
            
            if amount_col is not None: result['Amount'] = safe_convert_number(summary_row.iloc[amount_col])
            
            if left_col is not None: result['Left'] = safe_convert_number(summary_row.iloc[left_col])
            
            # Fallback hardcoded logic yang lebih sederhana
            if result['Amount'] is None and num_cols >= 10:
                # Coba ambil kolom-kolom ujung kanan yang biasanya berisi Amount/Sisa
                result['Amount'] = safe_convert_number(summary_row.iloc[-3]) # 3 dari kanan
                result['Left'] = safe_convert_number(summary_row.iloc[-1])   # Paling kanan
        
        else:
            return None, f"⚠️ Sheet '{sheet_name}': Baris summary/total tidak ditemukan"
            
        return result, f"✅ {nama_promo[:50]}"
        
    except Exception as e:
        return None, f"❌ Sheet '{sheet_name}': Error - {str(e)[:50]}"

def generate_summary(uploaded_file):
    try:
        xl = pd.ExcelFile(uploaded_file)
    except Exception as e:
        st.error(f"❌ Error membaca file: {str(e)}")
        return None, []
    
    sheet_names = xl.sheet_names
    results = []
    processed_sheets = []
    skipped_sheets = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, sheet_name in enumerate(sheet_names):
        try:
            # Baca sheet tanpa header dulu untuk scanning fleksibel
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
            
            if df.empty:
                continue
            
            result, message = process_sheet_robust(df, sheet_name)
            
            # Simpan hasil jika valid (minimal ada Nama Promo atau Angka)
            if result and (result['Nama Promo'] or result['All Count']):
                result['No.'] = len(results) + 1
                results.append(result)
                processed_sheets.append(message)
            else:
                skipped_sheets.append(message)
                
        except Exception as e:
            skipped_sheets.append(f"❌ Sheet '{sheet_name}': {str(e)}")
        
        # Update progress
        if idx % 10 == 0: # Update setiap 10 sheet agar tidak lag
            progress_bar.progress((idx + 1) / len(sheet_names))
            status_text.text(f"Memproses sheet {idx + 1} dari {len(sheet_names)}...")
    
    progress_bar.empty()
    status_text.empty()
    
    all_logs = processed_sheets + skipped_sheets
    
    if len(results) == 0:
        return None, all_logs
    
    summary_df = pd.DataFrame(results)
    
    cols = ['No.', 'Nama Promo', 'Mekanisme Promo', 'Periode Promo', 
            'All Count', 'All Claim', 'Sales Amount', 'Amount', 'Left']
    
    # Pastikan semua kolom ada
    for col in cols:
        if col not in summary_df.columns:
            summary_df[col] = None
            
    return summary_df[cols], all_logs

def format_preview_display(df):
    display_df = df.copy()
    numeric_cols = ['All Count', 'All Claim', 'Sales Amount', 'Amount', 'Left']
    
    for col in numeric_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(
                lambda x: '{:,.0f}'.format(x) if pd.notna(x) and x is not None else ''
            )
    return display_df

def save_summary_to_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['No.', 'Nama Promo', 'Mekanisme Promo', 'Periode Promo', 
               'All Count', 'All Claim', 'Sales Amount', 'Amount', 'Left']
    
    # Write Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write Data
    for row_idx, row_data in df.iterrows():
        excel_row = row_idx + 2
        for col_idx, header in enumerate(headers):
            value = row_data.get(header, '')
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            
            # Format Angka
            if header in ['All Count', 'All Claim', 'Sales Amount', 'Amount', 'Left']:
                if pd.notna(value) and value is not None:
                    try:
                        cell.value = int(float(value))
                        cell.number_format = '#,##0'
                    except:
                        cell.value = str(value)
                else:
                    cell.value = ''
            else:
                cell.value = str(value) if pd.notna(value) else ''
            
            cell.border = thin_border
            
            if col_idx == 0: cell.alignment = Alignment(horizontal="center")
            elif col_idx > 3: cell.alignment = Alignment(horizontal="right")
            else: cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            if row_idx % 2 == 1: cell.fill = alt_fill
            
    # Layout Adjustment
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# TAMPILAN APLIKASI
# ============================================================

st.title("🚀 Otomasi Summary Event Promo v5.1")
st.markdown("---")

with st.sidebar:
    st.header("📌 Panduan")
    st.info("Versi ini telah diperbarui untuk membaca lebih banyak format sheet dan membersihkan tanggal/angka secara otomatis.")

col1, col2 = st.columns([2, 1])
with col1:
    st.header("📤 Upload File Excel")
    uploaded_file = st.file_uploader("Pilih file Excel mentah", type=['xlsx', 'xls'])

if uploaded_file is not None:
    st.success(f"File: {uploaded_file.name}")
    if st.button("🚀 Proses File", type="primary", use_container_width=True):
        with st.spinner("Sedang memproses..."):
            summary_df, logs = generate_summary(uploaded_file)
            
            if summary_df is not None:
                st.session_state['data'] = summary_df
                st.session_state['logs'] = logs
                st.session_state['filename'] = uploaded_file.name
                st.success(f"Selesai! {len(summary_df)} baris data berhasil diekstrak.")
            else:
                st.error("Gagal memproses file.")

if 'data' in st.session_state:
    df = st.session_state['data']
    
    st.markdown("---")
    st.subheader("📊 Preview Hasil")
    st.dataframe(format_preview_display(df), use_container_width=True, hide_index=True)
    
    excel_data = save_summary_to_excel(df)
    filename = f"Summary_{st.session_state['filename']}"
    
    st.download_button("📥 Download Excel", excel_data, file_name=filename, 
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                       type="primary", use_container_width=True)

    with st.expander("Log Proses"):
        for log in st.session_state.get('logs', []):
            st.text(log)
